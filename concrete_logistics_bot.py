"""
Concrete Logistics Telegram Bot
Production | PostgreSQL | Render Web Service
Admin:  https://t.me/MaterialConcretebot?start=Admin
Worker: https://t.me/MaterialConcretebot?start=Worker
"""

import os, logging, io, threading, time, requests
from http.server import HTTPServer, BaseHTTPRequestHandler
from datetime import datetime, timedelta, date
from typing import Optional

import psycopg2, psycopg2.extras, psycopg2.pool
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, ContextTypes, ConversationHandler, filters,
)

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
BOT_TOKEN     = os.environ["BOT_TOKEN"]
DATABASE_URL  = os.environ["DATABASE_URL"]
BOT_USERNAME  = "MaterialConcretebot"
ADMIN_SECRET  = "Admin"
WORKER_SECRET = "Worker"
PORT          = int(os.environ.get("PORT", 10000))

CONCRETE_GRADES = [
    "C5","C8","C10","C12","C15","C16","C20",
    "C25","C30","C35","C40","C45","C50"
]

logging.basicConfig(format="%(asctime)s %(levelname)s %(message)s", level=logging.WARNING)
log = logging.getLogger(__name__)

# Conversation states — no overlaps
(ASK_JOB, ASK_GRADE, ASK_PLATE, ASK_PLATE_MANUAL,
 ASK_VOLUME, CONFIRM_TRIP) = range(6)
ASK_NEW_JOB_NAME = 10
ASK_NEW_TRUCK    = 20


# ─────────────────────────────────────────────
# HEALTH SERVER  (only used in polling/local mode)
# ─────────────────────────────────────────────
class _Health(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200); self.end_headers()
        self.wfile.write(b"OK")
    def log_message(self, *a): pass

def _start_health():
    HTTPServer(("0.0.0.0", PORT), _Health).serve_forever()


# ─────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────
_pool: Optional[psycopg2.pool.ThreadedConnectionPool] = None

def get_pool():
    global _pool
    if _pool is None or _pool.closed:
        _pool = psycopg2.pool.ThreadedConnectionPool(1, 10, DATABASE_URL)
    return _pool

def db(sql, params=(), *, one=False, many=False, rc=False):
    conn = get_pool().getconn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            conn.commit()
            if one:  return dict(cur.fetchone()) if cur.rowcount else None
            if many: return [dict(r) for r in cur.fetchall()]
            if rc:   return cur.rowcount
    except Exception:
        conn.rollback(); raise
    finally:
        get_pool().putconn(conn)

def init_db():
    conn = get_pool().getconn()
    try:
        with conn.cursor() as cur:
            # Create tables
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    user_id   BIGINT PRIMARY KEY,
                    username  TEXT,
                    role      TEXT NOT NULL DEFAULT 'worker',
                    joined_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                );
                CREATE TABLE IF NOT EXISTS jobs (
                    id         SERIAL PRIMARY KEY,
                    name       TEXT NOT NULL,
                    status     TEXT NOT NULL DEFAULT 'active',
                    created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    updated_at TIMESTAMPTZ
                );
                CREATE TABLE IF NOT EXISTS trucks (
                    id       SERIAL PRIMARY KEY,
                    plate    TEXT NOT NULL UNIQUE,
                    added_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                );
                CREATE TABLE IF NOT EXISTS trips (
                    id          SERIAL PRIMARY KEY,
                    user_id     BIGINT NOT NULL,
                    job_name    TEXT NOT NULL,
                    truck_plate TEXT NOT NULL,
                    volume      REAL NOT NULL
                );
            """)
            conn.commit()
            # Safe migrations
            cur.execute("""
                DO $$ BEGIN
                    IF EXISTS (SELECT 1 FROM information_schema.columns
                               WHERE table_name='trips' AND column_name='timestamp') THEN
                        ALTER TABLE trips RENAME COLUMN "timestamp" TO logged_at;
                    END IF;
                    IF NOT EXISTS (SELECT 1 FROM information_schema.columns
                                   WHERE table_name='trips' AND column_name='logged_at') THEN
                        ALTER TABLE trips ADD COLUMN logged_at TIMESTAMPTZ NOT NULL DEFAULT NOW();
                    END IF;
                    IF NOT EXISTS (SELECT 1 FROM information_schema.columns
                                   WHERE table_name='trips' AND column_name='concrete_grade') THEN
                        ALTER TABLE trips ADD COLUMN concrete_grade TEXT NOT NULL DEFAULT '';
                    END IF;
                END $$;
            """)
            conn.commit()
            # Indexes after columns exist
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_trips_logged ON trips(logged_at);
                CREATE INDEX IF NOT EXISTS idx_trips_job    ON trips(job_name);
                CREATE INDEX IF NOT EXISTS idx_jobs_status  ON jobs(status);
            """)
            conn.commit()
        log.warning("DB ready.")
    finally:
        get_pool().putconn(conn)


# ─────────────────────────────────────────────
# CACHE
# ─────────────────────────────────────────────
_cache: dict = {}
def cget(k): return _cache.get(k)
def cset(k, v): _cache[k] = v
def cdel(*ks): [_cache.pop(k, None) for k in ks]
def cdel_prefix(*ps):
    for k in list(_cache):
        if any(k.startswith(p) for p in ps): del _cache[k]


# ─────────────────────────────────────────────
# DATA LAYER
# ─────────────────────────────────────────────
def register_user(uid, uname, role):
    db("""INSERT INTO users (user_id,username,role,joined_at) VALUES (%s,%s,%s,NOW())
          ON CONFLICT(user_id) DO UPDATE SET role=EXCLUDED.role,username=EXCLUDED.username""",
       (uid, uname or "", role))
    cdel(f"role_{uid}")

def get_role(uid) -> Optional[str]:
    k = f"role_{uid}"
    v = cget(k)
    if v is not None: return v or None
    row = db("SELECT role FROM users WHERE user_id=%s", (uid,), one=True)
    r = row["role"] if row else None
    cset(k, r or ""); return r

def is_admin(uid): return get_role(uid) == "admin"

def get_trucks():
    if cget("trucks") is not None: return cget("trucks")
    rows = db("SELECT * FROM trucks ORDER BY plate", many=True) or []
    cset("trucks", rows); return rows

def add_truck(plate):
    try:
        db("INSERT INTO trucks (plate, added_at) VALUES (%s, NOW())", (plate.upper().strip(),))
        cdel("trucks"); return True
    except: return False

def clear_trucks():
    db("DELETE FROM trucks"); cdel("trucks")

def get_jobs(status=None):
    k = f"jobs_{status}"
    if cget(k) is not None: return cget(k)
    rows = (db("SELECT * FROM jobs WHERE status=%s ORDER BY created_at DESC", (status,), many=True)
            if status else db("SELECT * FROM jobs ORDER BY created_at DESC", many=True)) or []
    cset(k, rows); return rows

def get_job(jid):
    return db("SELECT * FROM jobs WHERE id=%s", (jid,), one=True)

def add_job(name):
    db("INSERT INTO jobs (name, status, created_at) VALUES (%s, 'active', NOW())", (name,))
    cdel("jobs_active","jobs_None")

def set_job_status(jid, status):
    db("UPDATE jobs SET status=%s,updated_at=NOW() WHERE id=%s", (status, jid))
    cdel("jobs_active","jobs_None")

def save_trip(uid, job, grade, plate, vol):
    db("INSERT INTO trips (user_id,job_name,concrete_grade,truck_plate,volume,logged_at) VALUES (%s,%s,%s,%s,%s,NOW())",
       (uid, job, grade, plate, vol))
    cdel_prefix("trips_","summary_","breakdown_")

def _range(period):
    t = date.today()
    if period == "daily":  return "logged_at::date=%s", (t,)
    if period == "weekly":
        m = t - timedelta(days=t.weekday())
        return "logged_at::date>=%s", (m,)
    return "logged_at::date>=%s", (t.replace(day=1),)

def fetch_trips(period):
    k = f"trips_{period}_{date.today()}"
    if cget(k): return cget(k)
    wh, p = _range(period)
    rows = db(f"SELECT * FROM trips WHERE {wh} ORDER BY logged_at DESC", p, many=True) or []
    cset(k, rows); return rows

def delete_trips(period):
    if period == "all": n = db("DELETE FROM trips", rc=True)
    else:
        wh, p = _range(period)
        n = db(f"DELETE FROM trips WHERE {wh}", p, rc=True)
    cdel_prefix("trips_","summary_","breakdown_")
    return n or 0

def job_summary(name):
    k = f"summary_{name}"
    if cget(k): return cget(k)
    row = db("SELECT COUNT(*) AS trips, COALESCE(SUM(volume),0) AS vol FROM trips WHERE job_name=%s",
             (name,), one=True) or {"trips":0,"vol":0}
    cset(k, row); return row

def grade_breakdown(name):
    k = f"breakdown_{name}"
    if cget(k) is not None: return cget(k)
    rows = db("""SELECT concrete_grade, COALESCE(SUM(volume),0) AS vol, COUNT(*) AS trips
                 FROM trips WHERE job_name=%s AND concrete_grade!=''
                 GROUP BY concrete_grade ORDER BY vol DESC""",
              (name,), many=True) or []
    cset(k, rows); return rows



def get_all_users():
    return db("SELECT * FROM users ORDER BY joined_at DESC", many=True) or []

def set_user_role(uid, role):
    db("UPDATE users SET role=%s WHERE user_id=%s", (role, uid))
    cdel(f"role_{uid}")

# ─────────────────────────────────────────────
# REPORTS
# ─────────────────────────────────────────────
def plabel(p):
    t = date.today()
    if p=="daily":  return f"Daily — {t.strftime('%b %d, %Y')}"
    if p=="weekly":
        m = t-timedelta(days=t.weekday())
        return f"Week {m.strftime('%b %d')}–{(m+timedelta(6)).strftime('%b %d, %Y')}"
    return f"Month of {t.strftime('%B %Y')}"

def text_report(period):
    trips = fetch_trips(period)
    if not trips: return f"📋 *{plabel(period)}*\n\n_No trips recorded._"
    jobs:dict={}; trucks:dict={}; grades:dict={}
    for t in trips:
        j=t["job_name"]; p=t["truck_plate"]
        g=t.get("concrete_grade") or "N/A"; v=float(t["volume"])
        for d,key in [(jobs,j),(trucks,p),(grades,g)]:
            if key not in d: d[key]=[0.0,0]
            d[key][0]+=v; d[key][1]+=1
    tv=sum(x[0] for x in jobs.values())
    tt=sum(x[1] for x in jobs.values())
    medals=["🥇","🥈","🥉"]
    jl="\n".join(f"  {i+1}. {n}: *{d[0]:.1f} m³* ({d[1]} trips)"
                 for i,(n,d) in enumerate(sorted(jobs.items(),key=lambda x:-x[1][0])[:5]))
    gl="\n".join(f"  🧱 {g}: *{d[0]:.1f} m³* ({d[1]} trips)"
                 for g,d in sorted(grades.items(),key=lambda x:-x[1][0]))
    tl="\n".join(f"  {medals[i] if i<3 else '▪️'} {pl}: *{d[0]:.1f} m³* ({d[1]} trips)"
                 for i,(pl,d) in enumerate(sorted(trucks.items(),key=lambda x:-x[1][0])))
    return (f"📋 *{plabel(period)}*\n━━━━━━━━━━━━━━━━━━━━\n\n"
            f"📅 *Jobs*\n{jl}\n  ───────\n  *{tv:.1f} m³ | {tt} trips*\n\n"
            f"🧱 *Concrete Grades*\n{gl}\n\n"
            f"🚛 *Trucks*\n{tl}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n_Generated {datetime.now().strftime('%H:%M %d %b %Y')}_")

def excel_report(period):
    trips = fetch_trips(period)
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title="Trips"
    hf=Font(bold=True,color="FFFFFF"); hfill=PatternFill("solid",fgColor="1F4E79")
    for c,h in enumerate(["#","Logged At","Job","Grade","Truck","m³"],1):
        cell=ws.cell(1,c,h); cell.font=hf; cell.fill=hfill
        cell.alignment=Alignment(horizontal="center")
    for i,t in enumerate(trips,2):
        ws.cell(i,1,i-1); ws.cell(i,2,str(t.get("logged_at",""))[:16])
        ws.cell(i,3,t["job_name"]); ws.cell(i,4,t.get("concrete_grade") or "N/A")
        ws.cell(i,5,t["truck_plate"]); ws.cell(i,6,float(t["volume"]))
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=20
    def _sh(title,hdrs,data):
        s=wb.create_sheet(title); s.append(hdrs)
        for c in s[1]: c.font=Font(bold=True)
        for row in data: s.append(row)
    jt:dict={}
    for t in trips:
        j=t["job_name"]
        if j not in jt: jt[j]=[0.0,0]
        jt[j][0]+=float(t["volume"]); jt[j][1]+=1
    _sh("Jobs",["Job","m³","Trips"],[[n,round(v,2),tr] for n,(v,tr) in sorted(jt.items(),key=lambda x:-x[1][0])])
    gt:dict={}
    for t in trips:
        g=t.get("concrete_grade") or "N/A"
        if g not in gt: gt[g]=[0.0,0]
        gt[g][0]+=float(t["volume"]); gt[g][1]+=1
    _sh("Grades",["Grade","m³","Trips"],[[g,round(v,2),tr] for g,(v,tr) in sorted(gt.items(),key=lambda x:-x[1][0])])
    tt2:dict={}
    for t in trips:
        p=t["truck_plate"]
        if p not in tt2: tt2[p]=[0.0,0]
        tt2[p][0]+=float(t["volume"]); tt2[p][1]+=1
    _sh("Trucks",["Rank","Truck","m³","Trips"],
        [[i,pl,round(v,2),tr] for i,(pl,(v,tr)) in enumerate(sorted(tt2.items(),key=lambda x:-x[1][0]),1)])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf


# ─────────────────────────────────────────────
# KEYBOARDS
# ─────────────────────────────────────────────
def kb_back():
    return InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back",callback_data="back_main")]])

def kb_main(uid):
    admin=is_admin(uid); kb=[]
    if admin:
        kb+=[[InlineKeyboardButton("🆕 Add Job",callback_data="add_job")],
             [InlineKeyboardButton("🚛 Manage Trucks",callback_data="manage_trucks")]]
    kb+=[[InlineKeyboardButton("➕ Log New Trip",callback_data="log_trip")],
         [InlineKeyboardButton("🏗️ Job Status",callback_data="job_status")],
         [InlineKeyboardButton("📊 View Reports",callback_data="menu_reports")],
         [InlineKeyboardButton("📥 Export Excel",callback_data="menu_export")]]
    if admin:
        kb+=[[InlineKeyboardButton("🗑️ Delete Reports",callback_data="delete_reports_menu")],
             [InlineKeyboardButton("✅ Complete Job",callback_data="complete_job")],
             [InlineKeyboardButton("❌ Cancel Job",callback_data="cancel_job")],
             [InlineKeyboardButton("👥 Manage Users",callback_data="manage_users")]]
    return InlineKeyboardMarkup(kb)

def kb_trucks():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Add Truck",callback_data="add_truck")],
        [InlineKeyboardButton("🗑️ Clear All Trucks",callback_data="clear_trucks")],
        [InlineKeyboardButton("⬅️ Back",callback_data="back_main")]])

def kb_grades():
    rows=[]; row=[]
    for i,g in enumerate(CONCRETE_GRADES,1):
        row.append(InlineKeyboardButton(g,callback_data=f"grade_{g}"))
        if i%3==0: rows.append(row); row=[]
    if row: rows.append(row)
    return InlineKeyboardMarkup(rows)

def kb_confirm():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Yes, correct",callback_data="plate_yes")],
        [InlineKeyboardButton("🔄 Change truck",callback_data="plate_no")]])

def kb_trip_done():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Log Another Trip",callback_data="log_trip")],
        [InlineKeyboardButton("📊 View Reports",callback_data="menu_reports")],
        [InlineKeyboardButton("🏠 Main Menu",callback_data="back_main")]])


# ─────────────────────────────────────────────
# HANDLERS — /start & navigation
# ─────────────────────────────────────────────
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u=update.effective_user; args=context.args or []
    if args:
        if args[0]==ADMIN_SECRET:
            register_user(u.id,u.username or "","admin")
            await update.message.reply_text(
                f"👑 *Welcome Admin {u.first_name}!*\n\n"
                f"Admin:  `https://t.me/{BOT_USERNAME}?start=Admin`\n"
                f"Worker: `https://t.me/{BOT_USERNAME}?start=Worker`",
                parse_mode="Markdown",reply_markup=kb_main(u.id))
        elif args[0]==WORKER_SECRET:
            register_user(u.id,u.username or "","worker")
            await update.message.reply_text(
                f"👷 *Welcome {u.first_name}!*\n\nWorker access granted.",
                parse_mode="Markdown",reply_markup=kb_main(u.id))
        else:
            await update.message.reply_text("❌ Invalid link.")
        return
    role=get_role(u.id)
    if role:
        lbl="👑 Admin" if role=="admin" else "👷 Worker"
        await update.message.reply_text(
            f"🏗️ *Concrete Logistics*\n_{lbl}_\n\nWelcome back, {u.first_name}!",
            parse_mode="Markdown",reply_markup=kb_main(u.id))
    else:
        await update.message.reply_text("🚫 Use your access link to register.")

async def cb_back_main(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    uid=q.from_user.id; role=get_role(uid)
    if not role: await q.edit_message_text("🚫 Access denied."); return
    lbl="👑 Admin" if role=="admin" else "👷 Worker"
    await q.edit_message_text(
        f"🏗️ *Concrete Logistics*\n_{lbl}_\n\nWhat would you like to do?",
        parse_mode="Markdown",reply_markup=kb_main(uid))

async def cb_noop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()


# ─────────────────────────────────────────────
# JOB STATUS
# ─────────────────────────────────────────────
async def cb_job_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    if not get_role(q.from_user.id):
        await q.edit_message_text("🚫 Access denied.",reply_markup=kb_back()); return
    jobs=get_jobs(status="active")
    if not jobs:
        await q.edit_message_text("🏗️ *Job Status*\n\n_No active jobs._",
            parse_mode="Markdown",reply_markup=kb_back()); return
    lines=[]
    for j in jobs:
        s=job_summary(j["name"]); bd=grade_breakdown(j["name"])
        block=(f"🟢 *{j['name']}*\n"
               f"   📦 *{float(s['vol']):.1f} m³* | *{s['trips']} trips*")
        if bd:
            block+="\n   🧱 *Grades:*\n"+"\n".join(
                f"      ▪️ {r['concrete_grade']}: *{float(r['vol']):.1f} m³* ({r['trips']} trips)"
                for r in bd)
        lines.append(block)
    await q.edit_message_text(
        "🏗️ *Active Job Sites*\n━━━━━━━━━━━━━━━━━━━━\n\n"+"\n\n".join(lines),
        parse_mode="Markdown",reply_markup=kb_back())


# ─────────────────────────────────────────────
# MANAGE TRUCKS
# ─────────────────────────────────────────────
async def cb_manage_trucks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    if not is_admin(q.from_user.id):
        await q.edit_message_text("🚫 Admins only.",reply_markup=kb_back()); return
    trucks=get_trucks()
    txt=("🚛 *Registered Trucks*\n━━━━━━━━━━━━━━━━━━━━\n\n"+
         "\n".join(f"🚛 {t['plate']}" for t in trucks)
         if trucks else "🚛 *Trucks*\n\n_No trucks yet._")
    await q.edit_message_text(txt,parse_mode="Markdown",reply_markup=kb_trucks())

async def cb_clear_trucks_ask(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    if not is_admin(q.from_user.id): return
    trucks=get_trucks()
    if not trucks:
        await q.edit_message_text("No trucks to clear.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back",callback_data="manage_trucks")]])); return
    await q.edit_message_text(
        f"⚠️ *Clear all {len(trucks)} trucks?*\n\n"+"\n".join(f"🚛 {t['plate']}" for t in trucks),
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Yes, clear all",callback_data="clear_trucks_yes")],
            [InlineKeyboardButton("❌ Cancel",callback_data="manage_trucks")]]))

async def cb_clear_trucks_do(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    if not is_admin(q.from_user.id): return
    clear_trucks()
    await q.edit_message_text("🗑️ All trucks cleared.",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🚛 Manage Trucks",callback_data="manage_trucks")],
            [InlineKeyboardButton("🏠 Main Menu",callback_data="back_main")]]))

async def conv_truck_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    if not is_admin(q.from_user.id):
        await q.edit_message_text("🚫 Admins only.",reply_markup=kb_back())
        return ConversationHandler.END
    await q.message.reply_text("🚛 Enter truck plate number:")
    return ASK_NEW_TRUCK

async def conv_truck_save(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    plate=update.message.text.strip().upper()
    ok=add_truck(plate)
    msg=f"✅ Truck `{plate}` added!" if ok else f"⚠️ `{plate}` already exists."
    await update.message.reply_text(msg,parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("➕ Add Another",callback_data="add_truck")],
            [InlineKeyboardButton("🚛 Manage Trucks",callback_data="manage_trucks")],
            [InlineKeyboardButton("🏠 Main Menu",callback_data="back_main")]]))
    return ConversationHandler.END


# ─────────────────────────────────────────────
# ADD JOB
# ─────────────────────────────────────────────
async def conv_job_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    if not is_admin(q.from_user.id):
        await q.edit_message_text("🚫 Admins only.",reply_markup=kb_back())
        return ConversationHandler.END
    await q.message.reply_text("🆕 Enter job site name:")
    return ASK_NEW_JOB_NAME

async def conv_job_save(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    name=update.message.text.strip()
    add_job(name)
    await update.message.reply_text(f"✅ *Job '{name}' added!*",parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🆕 Add Another",callback_data="add_job")],
            [InlineKeyboardButton("🏠 Main Menu",callback_data="back_main")]]))
    return ConversationHandler.END


# ─────────────────────────────────────────────
# COMPLETE / CANCEL JOB
# ─────────────────────────────────────────────
async def cb_complete_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    if not is_admin(q.from_user.id):
        await q.edit_message_text("🚫 Admins only.",reply_markup=kb_back()); return
    jobs=get_jobs(status="active")
    if not jobs:
        await q.edit_message_text("_No active jobs._",parse_mode="Markdown",reply_markup=kb_back()); return
    kb=[[InlineKeyboardButton(f"🏗️ {j['name']}",callback_data=f"do_complete_{j['id']}")] for j in jobs]
    kb.append([InlineKeyboardButton("⬅️ Back",callback_data="back_main")])
    await q.edit_message_text("✅ *Select job to complete:*",parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(kb))

async def cb_do_complete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    jid=int(q.data.replace("do_complete_",""))
    j=get_job(jid)
    if j:
        set_job_status(jid,"completed")
        s=job_summary(j["name"])
        await q.edit_message_text(
            f"✅ *{j['name']}* completed!\n📦 {float(s['vol']):.1f} m³ | {s['trips']} trips",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Main Menu",callback_data="back_main")]]))

async def cb_cancel_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    if not is_admin(q.from_user.id):
        await q.edit_message_text("🚫 Admins only.",reply_markup=kb_back()); return
    jobs=get_jobs(status="active")
    if not jobs:
        await q.edit_message_text("_No active jobs._",parse_mode="Markdown",reply_markup=kb_back()); return
    kb=[[InlineKeyboardButton(f"🏗️ {j['name']}",callback_data=f"do_cancel_{j['id']}")] for j in jobs]
    kb.append([InlineKeyboardButton("⬅️ Back",callback_data="back_main")])
    await q.edit_message_text("❌ *Select job to cancel:*",parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(kb))

async def cb_do_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    jid=int(q.data.replace("do_cancel_",""))
    j=get_job(jid)
    if j:
        set_job_status(jid,"cancelled")
        await q.edit_message_text(f"❌ *{j['name']}* cancelled.",parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Main Menu",callback_data="back_main")]]))


# ─────────────────────────────────────────────
# DELETE REPORTS
# ─────────────────────────────────────────────
_PL={"daily":"Today","weekly":"This week","monthly":"This month","all":"ALL trips ever"}

async def cb_del_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    if not is_admin(q.from_user.id):
        await q.edit_message_text("🚫 Admins only.",reply_markup=kb_back()); return
    await q.edit_message_text("🗑️ *Delete Records*\n\n⚠️ This is permanent.",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🗑️ Today",callback_data="delask_daily")],
            [InlineKeyboardButton("🗑️ This Week",callback_data="delask_weekly")],
            [InlineKeyboardButton("🗑️ This Month",callback_data="delask_monthly")],
            [InlineKeyboardButton("⚠️ Delete ALL",callback_data="delask_all")],
            [InlineKeyboardButton("⬅️ Back",callback_data="back_main")]]))

async def cb_del_ask(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; period=q.data.replace("delask_",""); await q.answer()
    await q.edit_message_text(
        f"⚠️ Delete *{_PL[period]}*?\n\nThis cannot be undone.",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Yes, delete",callback_data=f"deldo_{period}")],
            [InlineKeyboardButton("❌ Cancel",callback_data="delete_reports_menu")]]))

async def cb_del_do(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; period=q.data.replace("deldo_",""); await q.answer()
    n=delete_trips(period)
    await q.edit_message_text(f"🗑️ Deleted *{n}* records.",parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🗑️ Delete More",callback_data="delete_reports_menu")],
            [InlineKeyboardButton("🏠 Main Menu",callback_data="back_main")]]))



def get_all_users():
    return db("SELECT * FROM users ORDER BY joined_at DESC", many=True) or []

def set_user_role(uid, role):
    db("UPDATE users SET role=%s WHERE user_id=%s", (role, uid))
    cdel(f"role_{uid}")

# ─────────────────────────────────────────────
# REPORTS / EXPORT
# ─────────────────────────────────────────────
async def cb_menu_reports(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    await q.edit_message_text("📊 *Select period:*",parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("📅 Today",callback_data="rep_daily")],
            [InlineKeyboardButton("🗓️ This Week",callback_data="rep_weekly")],
            [InlineKeyboardButton("📆 This Month",callback_data="rep_monthly")],
            [InlineKeyboardButton("⬅️ Back",callback_data="back_main")]]))

async def cb_menu_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    await q.edit_message_text("📥 *Export to Excel:*",parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("📅 Today",callback_data="exp_daily")],
            [InlineKeyboardButton("🗓️ This Week",callback_data="exp_weekly")],
            [InlineKeyboardButton("📆 This Month",callback_data="exp_monthly")],
            [InlineKeyboardButton("⬅️ Back",callback_data="back_main")]]))

async def cb_text_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; period=q.data.replace("rep_",""); await q.answer()
    await q.edit_message_text(text_report(period),parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back",callback_data="menu_reports")]]))

async def cb_excel_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; period=q.data.replace("exp_","")
    await q.answer("Generating…")
    buf=excel_report(period)
    await context.bot.send_document(
        chat_id=q.message.chat_id,
        document=InputFile(buf,filename=f"trips_{period}_{date.today()}.xlsx"),
        caption=f"📥 *{plabel(period)}*",parse_mode="Markdown")


# ─────────────────────────────────────────────
# LOG TRIP  Job→Grade→Truck→Confirm→Volume
# ─────────────────────────────────────────────
async def conv_trip_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; uid=q.from_user.id; await q.answer()
    if not get_role(uid):
        await q.edit_message_text("🚫 Access denied."); return ConversationHandler.END
    jobs=get_jobs(status="active")
    if not jobs:
        await q.message.reply_text("⚠️ No active jobs. Ask admin to add one.",
            reply_markup=kb_back()); return ConversationHandler.END
    kb=[[InlineKeyboardButton(f"🏗️ {j['name']}",callback_data=f"tj_{j['id']}")] for j in jobs]
    await q.message.reply_text("📍 *Step 1/4 — Select Job Site:*",
        parse_mode="Markdown",reply_markup=InlineKeyboardMarkup(kb))
    return ASK_JOB

async def conv_trip_job(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    jid=int(q.data.replace("tj_",""))
    j=get_job(jid); context.user_data["job"]=j["name"]
    await q.message.reply_text(
        f"🧱 *Step 2/4 — Select Concrete Grade:*\nJob: `{j['name']}`",
        parse_mode="Markdown",reply_markup=kb_grades())
    return ASK_GRADE

async def conv_trip_grade(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    grade=q.data.replace("grade_",""); context.user_data["grade"]=grade
    trucks=get_trucks()
    kb=[[InlineKeyboardButton(f"🚛 {t['plate']}",callback_data=f"tp_{t['plate']}")] for t in trucks]
    kb.append([InlineKeyboardButton("✏️ Type manually",callback_data="tp_manual")])
    await q.message.reply_text(
        f"🚛 *Step 3/4 — Select Truck:*\n"
        f"Job: `{context.user_data['job']}` | Grade: `{grade}`",
        parse_mode="Markdown",reply_markup=InlineKeyboardMarkup(kb))
    return ASK_PLATE

async def conv_trip_plate_list(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; plate=q.data.replace("tp_",""); await q.answer()
    if plate=="manual":
        await q.message.reply_text("✏️ Type the truck plate:"); return ASK_PLATE_MANUAL
    context.user_data["plate"]=plate
    await q.message.reply_text(
        f"⚠️ *Confirm:*\n🏗️ `{context.user_data['job']}`\n"
        f"🧱 `{context.user_data['grade']}`\n🚛 `{plate}`\n\nCorrect?",
        parse_mode="Markdown",reply_markup=kb_confirm())
    return CONFIRM_TRIP

async def conv_trip_plate_manual(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    plate=update.message.text.strip().upper(); context.user_data["plate"]=plate
    await update.message.reply_text(
        f"⚠️ *Confirm:*\n🏗️ `{context.user_data['job']}`\n"
        f"🧱 `{context.user_data['grade']}`\n🚛 `{plate}`\n\nCorrect?",
        parse_mode="Markdown",reply_markup=kb_confirm())
    return CONFIRM_TRIP

async def conv_trip_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    if q.data=="plate_no":
        trucks=get_trucks()
        kb=[[InlineKeyboardButton(f"🚛 {t['plate']}",callback_data=f"tp_{t['plate']}")] for t in trucks]
        kb.append([InlineKeyboardButton("✏️ Type manually",callback_data="tp_manual")])
        await q.message.reply_text("🚛 *Select Truck:*",
            parse_mode="Markdown",reply_markup=InlineKeyboardMarkup(kb))
        return ASK_PLATE
    await q.message.reply_text(
        f"📦 *Step 4/4 — Enter Volume (m³):*\n"
        f"🏗️ `{context.user_data['job']}` | 🧱 `{context.user_data['grade']}` | 🚛 `{context.user_data['plate']}`",
        parse_mode="Markdown")
    return ASK_VOLUME

async def conv_trip_volume(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        vol=float(update.message.text.strip())
        if vol<=0: raise ValueError
    except ValueError:
        await update.message.reply_text("❌ Enter a valid positive number (e.g. 7.5)")
        return ASK_VOLUME
    job=context.user_data["job"]; grade=context.user_data["grade"]; plate=context.user_data["plate"]
    save_trip(update.effective_user.id,job,grade,plate,vol)
    context.user_data.clear()
    await update.message.reply_text(
        f"✅ *Trip Saved!*\n"
        f"📍 `{job}`\n"
        f"🧱 `{grade}`  🚛 `{plate}`  📦 `{vol} m³`\n"
        f"🕐 `{datetime.now().strftime('%H:%M, %d %b %Y')}`",
        parse_mode="Markdown",reply_markup=kb_trip_done())
    return ConversationHandler.END

async def conv_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    await update.message.reply_text("❌ Cancelled.")
    return ConversationHandler.END



# ─────────────────────────────────────────────
# USER MANAGEMENT (admin)
# ─────────────────────────────────────────────
async def cb_manage_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    if not is_admin(q.from_user.id):
        await q.edit_message_text("🚫 Admins only.",reply_markup=kb_back()); return
    users=get_all_users()
    if not users:
        await q.edit_message_text("👥 *Users*\n\n_No users registered._",
            parse_mode="Markdown",reply_markup=kb_back()); return
    kb=[]
    for u in users:
        role_icon="👑" if u["role"]=="admin" else "👷"
        uname=u["username"] or f"id:{u['user_id']}"
        kb.append([InlineKeyboardButton(
            f"{role_icon} {uname} — {u['role']}",
            callback_data=f"userinfo_{u['user_id']}")])
    kb.append([InlineKeyboardButton("⬅️ Back",callback_data="back_main")])
    await q.edit_message_text("👥 *User Management*\n━━━━━━━━━━━━━━━━━━━━\n\nSelect a user to manage:",
        parse_mode="Markdown",reply_markup=InlineKeyboardMarkup(kb))

async def cb_user_info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    if not is_admin(q.from_user.id): return
    uid=int(q.data.replace("userinfo_",""))
    users=get_all_users()
    u=next((x for x in users if x["user_id"]==uid),None)
    if not u:
        await q.edit_message_text("User not found.",reply_markup=kb_back()); return
    role_icon="👑" if u["role"]=="admin" else "👷"
    uname=u["username"] or f"id:{u['user_id']}"
    # Toggle button
    new_role="worker" if u["role"]=="admin" else "admin"
    new_icon="👷" if new_role=="worker" else "👑"
    await q.edit_message_text(
        f"👤 *User Details*\n━━━━━━━━━━━━━━━━━━━━\n\n"
        f"Name: {uname}\n"
        f"Role: {role_icon} *{u['role']}*\n"
        f"ID: `{u['user_id']}`\n"
        f"Joined: {str(u['joined_at'])[:10]}",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(f"Change to {new_icon} {new_role}",
                callback_data=f"setrole_{uid}_{new_role}")],
            [InlineKeyboardButton("🗑️ Remove User",
                callback_data=f"removeuser_{uid}")],
            [InlineKeyboardButton("⬅️ Back",callback_data="manage_users")],
        ]))

async def cb_set_role(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    if not is_admin(q.from_user.id): return
    parts=q.data.replace("setrole_","").split("_")
    uid=int(parts[0]); new_role=parts[1]
    set_user_role(uid,new_role)
    icon="👑" if new_role=="admin" else "👷"
    await q.edit_message_text(
        f"✅ User role updated to {icon} *{new_role}*",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("👥 Back to Users",callback_data="manage_users")],
            [InlineKeyboardButton("🏠 Main Menu",callback_data="back_main")],
        ]))

async def cb_remove_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    if not is_admin(q.from_user.id): return
    uid=int(q.data.replace("removeuser_",""))
    users=get_all_users()
    u=next((x for x in users if x["user_id"]==uid),None)
    uname=u["username"] if u else str(uid)
    await q.edit_message_text(
        f"⚠️ Remove user *{uname}*?\n\nThey will lose access.",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Yes, remove",callback_data=f"confirmremove_{uid}")],
            [InlineKeyboardButton("❌ Cancel",callback_data=f"userinfo_{uid}")],
        ]))

async def cb_confirm_remove(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    if not is_admin(q.from_user.id): return
    uid=int(q.data.replace("confirmremove_",""))
    db("DELETE FROM users WHERE user_id=%s",(uid,))
    cdel(f"role_{uid}")
    await q.edit_message_text("🗑️ User removed.",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("👥 Back to Users",callback_data="manage_users")],
            [InlineKeyboardButton("🏠 Main Menu",callback_data="back_main")],
        ]))

# ─────────────────────────────────────────────
# CONFLICT KILLER
# ─────────────────────────────────────────────
def kill_webhook():
    """Delete webhook and wait for Telegram to release the session."""
    for attempt in range(5):
        try:
            r=requests.post(f"https://api.telegram.org/bot{BOT_TOKEN}/deleteWebhook",
                            json={"drop_pending_updates":True},timeout=10)
            log.warning(f"deleteWebhook attempt {attempt+1} → {r.json()}")
            if r.json().get("ok"): break
        except Exception as e:
            log.warning(f"deleteWebhook error: {e}")
        time.sleep(3)
    time.sleep(5)  # extra wait to let old session die


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    init_db()
    WEBHOOK_URL = os.environ.get("RENDER_EXTERNAL_URL","")
    if WEBHOOK_URL:
        # Delete any existing webhook/polling before setting new one
        kill_webhook()
    else:
        # local dev only
        threading.Thread(target=_start_health,daemon=True).start()
        kill_webhook()

    app=(Application.builder()
         .token(BOT_TOKEN)
         .concurrent_updates(False)
         .connection_pool_size(8)
         .build())

    add_job_conv=ConversationHandler(
        entry_points=[CallbackQueryHandler(conv_job_start,pattern="^add_job$")],
        states={ASK_NEW_JOB_NAME:[MessageHandler(filters.TEXT&~filters.COMMAND,conv_job_save)]},
        fallbacks=[CommandHandler("cancel",conv_cancel)],
        allow_reentry=True,
    )
    add_truck_conv=ConversationHandler(
        entry_points=[CallbackQueryHandler(conv_truck_start,pattern="^add_truck$")],
        states={ASK_NEW_TRUCK:[MessageHandler(filters.TEXT&~filters.COMMAND,conv_truck_save)]},
        fallbacks=[CommandHandler("cancel",conv_cancel)],
        allow_reentry=True,
    )
    log_trip_conv=ConversationHandler(
        entry_points=[CallbackQueryHandler(conv_trip_start,pattern="^log_trip$")],
        states={
            ASK_JOB:          [CallbackQueryHandler(conv_trip_job,         pattern="^tj_\\d+$")],
            ASK_GRADE:        [CallbackQueryHandler(conv_trip_grade,       pattern="^grade_")],
            ASK_PLATE:        [CallbackQueryHandler(conv_trip_plate_list,  pattern="^tp_")],
            ASK_PLATE_MANUAL: [MessageHandler(filters.TEXT&~filters.COMMAND,conv_trip_plate_manual)],
            CONFIRM_TRIP:     [CallbackQueryHandler(conv_trip_confirm,     pattern="^plate_(yes|no)$")],
            ASK_VOLUME:       [MessageHandler(filters.TEXT&~filters.COMMAND,conv_trip_volume)],
        },
        fallbacks=[CommandHandler("cancel",conv_cancel)],
        allow_reentry=True,
    )

    # Register conversation handlers first (group 0)
    app.add_handler(add_job_conv,   group=0)
    app.add_handler(add_truck_conv, group=0)
    app.add_handler(log_trip_conv,  group=0)

    # Then all other handlers (group 1)
    app.add_handler(CommandHandler("start",cmd_start),                                              group=1)
    app.add_handler(CallbackQueryHandler(cb_back_main,       pattern="^back_main$"),                group=1)
    app.add_handler(CallbackQueryHandler(cb_job_status,      pattern="^job_status$"),               group=1)
    app.add_handler(CallbackQueryHandler(cb_manage_trucks,   pattern="^manage_trucks$"),            group=1)
    app.add_handler(CallbackQueryHandler(cb_clear_trucks_ask,pattern="^clear_trucks$"),             group=1)
    app.add_handler(CallbackQueryHandler(cb_clear_trucks_do, pattern="^clear_trucks_yes$"),         group=1)
    app.add_handler(CallbackQueryHandler(cb_complete_menu,   pattern="^complete_job$"),             group=1)
    app.add_handler(CallbackQueryHandler(cb_cancel_menu,     pattern="^cancel_job$"),               group=1)
    app.add_handler(CallbackQueryHandler(cb_do_complete,     pattern="^do_complete_\\d+$"),         group=1)
    app.add_handler(CallbackQueryHandler(cb_do_cancel,       pattern="^do_cancel_\\d+$"),           group=1)
    app.add_handler(CallbackQueryHandler(cb_del_menu,        pattern="^delete_reports_menu$"),      group=1)
    app.add_handler(CallbackQueryHandler(cb_del_ask,         pattern="^delask_(daily|weekly|monthly|all)$"), group=1)
    app.add_handler(CallbackQueryHandler(cb_del_do,          pattern="^deldo_(daily|weekly|monthly|all)$"),  group=1)
    app.add_handler(CallbackQueryHandler(cb_menu_reports,    pattern="^menu_reports$"),             group=1)
    app.add_handler(CallbackQueryHandler(cb_menu_export,     pattern="^menu_export$"),              group=1)
    app.add_handler(CallbackQueryHandler(cb_text_report,     pattern="^rep_(daily|weekly|monthly)$"),group=1)
    app.add_handler(CallbackQueryHandler(cb_excel_report,    pattern="^exp_(daily|weekly|monthly)$"),group=1)
    app.add_handler(CallbackQueryHandler(cb_manage_users,   pattern="^manage_users$"),              group=1)
    app.add_handler(CallbackQueryHandler(cb_user_info,       pattern="^userinfo_\\d+$"),             group=1)
    app.add_handler(CallbackQueryHandler(cb_set_role,        pattern="^setrole_\\d+_(admin|worker)$"),group=1)
    app.add_handler(CallbackQueryHandler(cb_remove_user,     pattern="^removeuser_\\d+$"),           group=1)
    app.add_handler(CallbackQueryHandler(cb_confirm_remove,  pattern="^confirmremove_\\d+$"),        group=1)
    app.add_handler(CallbackQueryHandler(cb_noop,            pattern="^noop$"),                        group=1)

    async def error_handler(update, context):
        log.warning(f"Error: {context.error}")
    app.add_error_handler(error_handler)

    WEBHOOK_URL = os.environ.get("RENDER_EXTERNAL_URL", "")
    if WEBHOOK_URL:
        # Webhook mode — no conflicts, production ready
        log.warning(f"Starting webhook on {WEBHOOK_URL}")
        app.run_webhook(
            listen="0.0.0.0",
            port=PORT,
            webhook_url=f"{WEBHOOK_URL}/webhook",
            drop_pending_updates=True,
        )
    else:
        # Fallback to polling for local dev
        log.warning("Starting polling (local mode)…")
        app.run_polling(poll_interval=1.0, timeout=30, drop_pending_updates=True)

if __name__=="__main__":
    main()
